import openpyxl
import csv
import os

wb = openpyxl.load_workbook('Master_Priority_Agenda_Review.xlsx', data_only=True)
os.makedirs('data/csv', exist_ok=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Clean sheet name for filename
    safe_name = sheet_name.replace(' ', '_').replace('/', '_')
    csv_path = f'data/csv/{safe_name}.csv'
    
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
            
print("Conversion complete.")
